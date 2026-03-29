"""
report_generator.py
===================
Generates formatted Excel test execution reports using openpyxl and pandas.

Produces:
    - Summary sheet  : overall pass/fail counts, coverage %, timestamp
    - Results sheet  : per-test result with signal values, duration, DTC status
    - Coverage sheet : requirements vs test case mapping

Author : Rakesh Aragodi
"""

from __future__ import annotations
import logging
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

logger = logging.getLogger(__name__)

# ── Colour palette ─────────────────────────────────────────────────────
CLR = {
    "header_bg"  : "1F3864",   # dark navy
    "header_fg"  : "FFFFFF",
    "pass_bg"    : "C6EFCE",   # light green
    "pass_fg"    : "276221",
    "fail_bg"    : "FFC7CE",   # light red
    "fail_fg"    : "9C0006",
    "abort_bg"   : "FFEB9C",   # amber
    "abort_fg"   : "9C6500",
    "alt_row"    : "F2F7FF",   # very light blue for alternating rows
    "border"     : "CCCCCC",
}

THIN_BORDER = Border(
    left=Side(style="thin", color=CLR["border"]),
    right=Side(style="thin", color=CLR["border"]),
    top=Side(style="thin", color=CLR["border"]),
    bottom=Side(style="thin", color=CLR["border"]),
)


# ──────────────────────────────────────────────────────────────────────
# Data class
# ──────────────────────────────────────────────────────────────────────

@dataclass
class TestResultRow:
    """One row in the results sheet."""
    test_id:      str
    test_name:    str
    feature:      str                   # "BCM" | "UDS"
    status:       str                   # "PASS" | "FAIL" | "ABORTED"
    duration_ms:  int
    requirement:  str = ""              # Jira requirement key e.g. BCM-REQ-012
    xray_key:     str = ""              # Jira Xray test key
    signal_name:  str = ""
    signal_value: Optional[float] = None
    expected:     Optional[float] = None
    dtc_present:  bool = False
    notes:        str = ""


# ──────────────────────────────────────────────────────────────────────
# Report generator
# ──────────────────────────────────────────────────────────────────────

class ReportGenerator:
    """
    Generates a multi-sheet Excel test report.

    Usage:
        rg = ReportGenerator()
        rg.add_result(TestResultRow(...))
        rg.add_result(TestResultRow(...))
        path = rg.save("reports/")
    """

    def __init__(self, project: str = "BCM SDV", build: str = ""):
        self._project  = project
        self._build    = build or datetime.now().strftime("%Y%m%d_%H%M%S")
        self._results: list[TestResultRow] = []

    def add_result(self, result: TestResultRow) -> None:
        self._results.append(result)

    def add_results(self, results: list[TestResultRow]) -> None:
        self._results.extend(results)

    # ------------------------------------------------------------------
    # Public: save to file
    # ------------------------------------------------------------------

    def save(self, output_dir: str = "reports") -> Path:
        """
        Write the Excel report and return the file path.

        Args:
            output_dir: Directory to write the report into.

        Returns:
            Path to the generated .xlsx file.
        """
        Path(output_dir).mkdir(parents=True, exist_ok=True)
        filename = f"test_report_{self._build}.xlsx"
        filepath = Path(output_dir) / filename

        wb = openpyxl.Workbook()
        wb.remove(wb.active)   # remove default blank sheet

        self._build_summary_sheet(wb)
        self._build_results_sheet(wb)
        self._build_coverage_sheet(wb)

        wb.save(filepath)
        logger.info("Report saved: %s", filepath)
        return filepath

    # ------------------------------------------------------------------
    # Sheet 1 — Summary
    # ------------------------------------------------------------------

    def _build_summary_sheet(self, wb: openpyxl.Workbook) -> None:
        ws = wb.create_sheet("Summary")

        total   = len(self._results)
        passed  = sum(1 for r in self._results if r.status == "PASS")
        failed  = sum(1 for r in self._results if r.status == "FAIL")
        aborted = total - passed - failed
        pct     = round(passed / total * 100, 1) if total else 0

        # Title
        ws.merge_cells("A1:D1")
        cell = ws["A1"]
        cell.value     = f"{self._project} — Test Execution Report"
        cell.font      = Font(bold=True, size=14, color=CLR["header_fg"])
        cell.fill      = PatternFill("solid", fgColor=CLR["header_bg"])
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        # Meta info
        meta = [
            ("Build / Run ID",   self._build),
            ("Executed",         datetime.now().strftime("%d %b %Y  %H:%M")),
            ("Total tests",      total),
            ("Passed",           passed),
            ("Failed",           failed),
            ("Aborted / Skipped", aborted),
            ("Pass rate",        f"{pct}%"),
        ]
        for i, (label, value) in enumerate(meta, start=3):
            ws.cell(row=i, column=1, value=label).font = Font(bold=True)
            ws.cell(row=i, column=2, value=value)

        # Bar chart — PASS / FAIL / ABORTED
        ws["D3"] = "Status"
        ws["E3"] = "Count"
        ws["D4"] = "PASS";  ws["E4"] = passed
        ws["D5"] = "FAIL";  ws["E5"] = failed
        ws["D6"] = "ABORTED"; ws["E6"] = aborted

        chart = BarChart()
        chart.type  = "col"
        chart.title = "Result Distribution"
        chart.style = 10
        chart.y_axis.title = "Test count"
        chart.x_axis.title = "Status"

        data = Reference(ws, min_col=5, min_row=3, max_row=6)
        cats = Reference(ws, min_col=4, min_row=4, max_row=6)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        ws.add_chart(chart, "A12")

        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 24

    # ------------------------------------------------------------------
    # Sheet 2 — Results
    # ------------------------------------------------------------------

    def _build_results_sheet(self, wb: openpyxl.Workbook) -> None:
        ws = wb.create_sheet("Results")

        headers = [
            "Test ID", "Test Name", "Feature", "Status",
            "Duration (ms)", "Signal Name", "Signal Value",
            "Expected", "DTC Present", "Requirement", "Xray Key", "Notes"
        ]
        col_widths = [10, 36, 10, 10, 14, 22, 14, 10, 12, 16, 12, 30]

        # Header row
        for col, (header, width) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font      = Font(bold=True, color=CLR["header_fg"])
            cell.fill      = PatternFill("solid", fgColor=CLR["header_bg"])
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = THIN_BORDER
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.row_dimensions[1].height = 20
        ws.freeze_panes = "A2"

        # Data rows
        for row_idx, result in enumerate(self._results, start=2):
            row_data = [
                result.test_id,
                result.test_name,
                result.feature,
                result.status,
                result.duration_ms,
                result.signal_name,
                result.signal_value,
                result.expected,
                "YES" if result.dtc_present else "NO",
                result.requirement,
                result.xray_key,
                result.notes,
            ]

            # Alternating row background
            row_fill = PatternFill("solid", fgColor=CLR["alt_row"] if row_idx % 2 == 0 else "FFFFFF")

            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border    = THIN_BORDER
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.fill      = row_fill

            # Colour-code the Status cell (column 4)
            status_cell = ws.cell(row=row_idx, column=4)
            if result.status == "PASS":
                status_cell.fill = PatternFill("solid", fgColor=CLR["pass_bg"])
                status_cell.font = Font(bold=True, color=CLR["pass_fg"])
            elif result.status == "FAIL":
                status_cell.fill = PatternFill("solid", fgColor=CLR["fail_bg"])
                status_cell.font = Font(bold=True, color=CLR["fail_fg"])
            else:
                status_cell.fill = PatternFill("solid", fgColor=CLR["abort_bg"])
                status_cell.font = Font(bold=True, color=CLR["abort_fg"])

    # ------------------------------------------------------------------
    # Sheet 3 — Coverage (requirements traceability)
    # ------------------------------------------------------------------

    def _build_coverage_sheet(self, wb: openpyxl.Workbook) -> None:
        ws = wb.create_sheet("Coverage")

        # Build a pandas DataFrame for coverage analysis
        rows = [
            {
                "Requirement"  : r.requirement or "N/A",
                "Test ID"      : r.test_id,
                "Test Name"    : r.test_name,
                "Status"       : r.status,
                "Xray Key"     : r.xray_key,
            }
            for r in self._results
        ]
        df = pd.DataFrame(rows)

        # Summary per requirement
        if not df.empty and "Requirement" in df.columns:
            summary = (
                df.groupby("Requirement")["Status"]
                .agg(
                    Total="count",
                    Passed=lambda x: (x == "PASS").sum(),
                    Failed=lambda x: (x == "FAIL").sum(),
                )
                .reset_index()
            )
            summary["Coverage %"] = (summary["Passed"] / summary["Total"] * 100).round(1)
        else:
            summary = pd.DataFrame(columns=["Requirement", "Total", "Passed", "Failed", "Coverage %"])

        headers = list(summary.columns)
        col_widths = [20, 8, 8, 8, 14]

        for col, (header, width) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font      = Font(bold=True, color=CLR["header_fg"])
            cell.fill      = PatternFill("solid", fgColor=CLR["header_bg"])
            cell.alignment = Alignment(horizontal="center")
            cell.border    = THIN_BORDER
            ws.column_dimensions[get_column_letter(col)].width = width

        for row_idx, row in enumerate(summary.itertuples(index=False), start=2):
            for col, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border    = THIN_BORDER
                cell.alignment = Alignment(horizontal="center")

            # Colour-code coverage %
            cov_cell = ws.cell(row=row_idx, column=5)
            cov = row[-1]
            if cov == 100:
                cov_cell.fill = PatternFill("solid", fgColor=CLR["pass_bg"])
                cov_cell.font = Font(bold=True, color=CLR["pass_fg"])
            elif cov >= 50:
                cov_cell.fill = PatternFill("solid", fgColor=CLR["abort_bg"])
                cov_cell.font = Font(bold=True, color=CLR["abort_fg"])
            else:
                cov_cell.fill = PatternFill("solid", fgColor=CLR["fail_bg"])
                cov_cell.font = Font(bold=True, color=CLR["fail_fg"])
